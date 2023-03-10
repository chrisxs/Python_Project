import re
import requests
from bs4 import BeautifulSoup
import os
import openpyxl
from datetime import datetime


# 创建一个images文件夹，如果该文件夹不存在
if not os.path.exists('images'):
    os.mkdir('images')

# 用户输入要爬取的网站URL
url = input('请输入要爬取的网站URL：')

# 发送请求并获取响应内容
response = requests.get(url)

# 使用BeautifulSoup解析HTML内容
soup = BeautifulSoup(response.content, 'html.parser')

# 获取网页标题
title = soup.title.string

# 获取所有图片链接
img_links = []
for img in soup.find_all('img'):
    src = img.get('src')
    if src is not None:
        # 使用正则表达式判断链接是否以http或https开头
        if re.match(r'^https?://', src):
            img_links.append(src)
        else:
            # 如果链接以//开头，则加上http:前缀
            if src.startswith('//'):
                img_links.append('http:' + src)

# 下载图片并保存到images文件夹
for i, img_url in enumerate(img_links):
    try:
        response = requests.get(img_url)
        with open(f'images/img_{i}.jpg', 'wb') as f:
            f.write(response.content)
    except Exception as e:
        print(f'下载图片 {img_url} 时发生异常: {e}')

# 获取所有链接和链接标题
links = []
for link in soup.find_all('a'):
    href = link.get('href')
    title = link.string
    if href is not None:
        # 使用正则表达式判断链接是否以http或https开头
        if re.match(r'^https?://', href):
            links.append({'title': title, 'href': href})
        else:
            # 如果链接以//开头，则加上http:前缀
            if href.startswith('//'):
                links.append({'title': title, 'href': 'http:' + href})
            # 如果链接以/开头，则加上网站域名前缀
            elif href.startswith('/'):
                links.append({'title': title, 'href': url + href})
            # 其他情况，直接使用链接地址
            else:
                links.append({'title': title, 'href': href})

# 创建工作簿和工作表
wb = openpyxl.Workbook()
ws = wb.active

# 写入标题和表头
ws['A1'] = '网页标题'
ws['B1'] = '图片链接'
ws['C1'] = '链接标题'
ws['D1'] = '链接地址'

# 写入数据
for i in range(len(img_links)):
    ws.cell(row=i+2, column=2, value=img_links[i])
for i in range(len(links)):
    if links[i]['title'] is not None:
        ws.cell(row=i+2, column=3, value=links[i]['title'])
    ws.cell(row=i+2, column=4, value=links[i]['href'])
ws['A2'] = title

# 保存工作簿
now = datetime.now()
filename = now.strftime('%Y-%m-%d %H-%M-%S') + '.xlsx'
wb.save(filename)

