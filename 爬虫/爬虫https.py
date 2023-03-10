import os
import re
import time
import requests
from datetime import datetime
from urllib.parse import urlparse
from bs4 import BeautifulSoup
from openpyxl import Workbook
from tqdm import tqdm

# 获取输入的网址123
url = input("滑稽爬虫HTTPS优化版\n请输入网址：https://")

# 判断协议类型
parsed_url = urlparse(url)
scheme = parsed_url.scheme
if not scheme:
    url = "https://" + url

# 获取网页内容
response = requests.get(url)
soup = BeautifulSoup(response.text, "html.parser")

# 获取所有标题
titles = []
for title in soup.find_all("title"):
    titles.append(title.string)
# print(f"标题：{titles}\n" )
print("\n标题：\n")
for titles in titles:
    print(titles)

# 获取所有链接
links = []
for link in soup.find_all("a"):
    href = link.get("href")
    if href and href.startswith("http"):
        links.append(href)
# print(f"链接：{links}")
print("\n链接:")
for link in links:
    print(link)

# 获取所有图片链接
img_links = []
for img in soup.find_all("img"):
    src = img.get("src")
    if src and src.startswith("http"):
        img_links.append(src)
# print(f"图片链接：{img_links}")
print("\n图片链接:")
for img_link in img_links:
    print(img_link)

# 创建xlsx文件
now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
workbook = Workbook()
worksheet = workbook.active
worksheet.title = "Links"
worksheet.cell(1, 1, value="标题")
worksheet.cell(1, 2, value="链接")
worksheet.cell(1, 3, value="图片链接")
for i, title in enumerate(titles):
    worksheet.cell(i+2, 1, value=title)
for i, link in enumerate(links):
    worksheet.cell(i+2, 2, value=link)
for i, img_link in enumerate(img_links):
    worksheet.cell(i+2, 3, value=img_link)
workbook.save(f"{now}.xlsx")
print(f"\n已导出xlsx文件：{now}.xlsx")

# 创建images文件夹
if not os.path.exists("images"):
    os.makedirs("images")

# 下载所有图片
for i, img_link in enumerate(tqdm(img_links, desc="下载图片")):
    try:
        response = requests.get(img_link)
        content_type = response.headers["Content-Type"]
        extension = re.findall(r"image/(.+)", content_type)[0]
        if not extension.startswith("jpeg") and not extension.startswith("png"):
            extension = "jpeg"
        filename = os.path.basename(urlparse(img_link).path)
        if not filename:
            filename = f"image_{i+1}.{extension}"
        else:
            filename = f"{filename}.{extension}"
        filename = re.sub(r'[\\/:*?"<>|\r\n]+', "_", filename)  # 替换非法字符
        filepath = os.path.join("images", filename)
        with open(filepath, "wb") as f:
            f.write(response.content)
    except Exception as e:
        print(f"下载图片{img_link}时出错\n{e}")
#     print("\n图片链接:")
# for img_link in img_links:
    # print(img_link)
print("\n所有图片已下载完毕")


# 等待用户手动关闭窗口
input("\n所有操作已完成，请按 Enter 键关闭窗口。")
